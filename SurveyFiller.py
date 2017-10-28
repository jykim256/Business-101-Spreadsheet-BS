import random
import openpyxl
import os
import time
import numpy as np

sample_size = 253
n_question = 29
n_answer = 7
XLName = 'Marketing.xlsx'
data = 'data.txt'


def genRow():
    row = [0 for i in range(n_answer)]
    for i in range(len(row) - 1):
        row[i] = random.randint(0, int((sample_size - sum(row)) / 1.6))
    row[len(row) - 1] = sample_size - sum(row)
    random.shuffle(row)
    return row


def assign(cell, value):
    cell.value = value


# if no data is given, will randomly make it up
if os.path.isfile(data):
    A = np.loadtxt(data)
else:
    A = np.array([genRow() for r in range(n_question)])
    
# initialize B array (responses)
B = np.array([[0 for c in range(n_question)] for r in range(sample_size)])
print(A)
print(B)

# fill array
for r in range(len(B)):
    for c in range(len(B[r])):
        answer = random.randint(0, n_answer - 1)
        while A[c][answer] < 1:
            answer = random.randint(0, n_answer - 1)
        B[r][c] = answer + 1
        A[c][answer] -= 1
print(A)
print(B)


try:
    book = openpyxl.load_workbook(filename=XLName)
    sheet = book.active
except:
    print('Book not found')
    book = openpyxl.Workbook()
    sheet = book.active

# Label survey #
for r in range(len(B)):
    sheet.cell(row=r + 2, column=1).value = 'S #: ' + str(r + 1)
# Label Question #
for c in range(len(B[0])):
    sheet.cell(row=1, column=c + 2).value = 'Q #: ' + str(c + 1)
# Transfer data from B array to sheet
for r in range(len(B)):
    for c in range(len(B[r])):
        sheet.cell(row=r + 2, column=c + 2).value = int(str(B[r][c]))

# Label question # for the count table
for c in range(len(B[0])):
    sheet.cell(row=1 + len(B) + 3, column=c + 2).value = 'Q #: ' + str(c + 1)
answer_choice = ['1 (Not at All)', '2 (A little)', '3 (Neutral)', '4 (Somewhat)',
                 '5 (Very)', '6 (Demographic only)', '7 (Demographic only)', 'TOTAL:']
for r in range(len(answer_choice)):
    sheet.cell(row=1 + len(B) + r + 3 + 1, column=1).value = answer_choice[r]

for r in range(len(answer_choice) - 1):
    for c in range(len(B[0])):
        char = chr(ord('B') + c)
        sheet.cell(row=1 + len(B) + 3 + 1 + r, column=c +
                   2).value = '=COUNTIF(' + char + '2:' + char + '254,' + str(r + 1) + ')'

for c in range(len(B[0])):
    char = chr(ord('B') + c)
    sheet.cell(row=1 + len(B) + 3 + len(answer_choice),
               column=c + 2).value = '=SUM()'

isOpen = True
while(isOpen):
    try:
        book.save(XLName)
        isOpen = False
    except:
        #os.system('taskkill /F /IM excel.exe')
        print('Close Excel Spreadsheet!')
        time.sleep(0.8)
print(XLName + ' updated.')
os.system(XLName)
